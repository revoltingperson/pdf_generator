import os
from pathlib import Path

from PyQt5 import uic
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import QWidget, QPushButton, QSpinBox, QLabel, QFileDialog, QComboBox, QTabWidget
from openpyxl import load_workbook, Workbook


class ExcelWindow(QWidget):
    opened_excel: Workbook
    window_signal = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        path = Path.cwd().parent.joinpath('ui', 'pdf_generator.ui')
        uic.loadUi(path, self)
        self.setWindowFlag(Qt.Window)
        self.choose_file: QPushButton = self.findChild(QPushButton, 'ChooseButton')
        self.choose_file.clicked.connect(self.open_file)
        self.col_start: QSpinBox = self.findChild(QSpinBox, 'Col_start')
        self.col_start.valueChanged.connect(self.spin_clicked)
        self.col_end: QSpinBox = self.findChild(QSpinBox, 'Col_end')
        self.col_end.valueChanged.connect(self.spin_clicked)
        self.row_start: QSpinBox = self.findChild(QSpinBox, 'Row_start')
        self.row_start.valueChanged.connect(self.spin_clicked)
        self.row_end: QSpinBox = self.findChild(QSpinBox, 'Row_end')
        self.row_end.valueChanged.connect(self.spin_clicked)
        self.generate = self.findChild(QPushButton, 'GenerateButton')
        self.generate.clicked.connect(self.generate_pdfs)
        self.message = self.findChild(QLabel, 'Output_message')
        self.filename: QLabel = self.findChild(QLabel, 'Filename')
        self.sheets: QComboBox = self.findChild(QComboBox, 'Sheets')
        self.tab_wid: QTabWidget = self.findChild(QTabWidget, 'tabWidget')
        self.data_preview = self.findChild(QLabel, 'Displayed_data')
        self.tab_wid.setTabEnabled(1, False)

    def open_file(self):
        file, _ = QFileDialog.getOpenFileNames(None, 'Open File', os.path.abspath(__name__),
                                               "Image Files(*.xlsx)")
        if file:
            head, file_name = os.path.split(file[0])
            self.filename.setText(file_name)
            self.opened_excel = load_workbook(file[0])
            names = self.opened_excel.sheetnames
            self.sheets.addItems(names)
            self.tab_wid.setTabEnabled(1, True)
            self.display_how_many()

    def extract_data_from_pdfs(self):
        if hasattr(self, 'opened_excel'):
            ready_to_use = []
            name = self.sheets.currentText()
            selected_ws = self.opened_excel[name]
            for row in range(self.row_start.value(), self.row_end.value()+1):
                string_from_one_row = []
                for col in range(self.col_start.value(), self.col_end.value()+1):
                    cell = selected_ws.cell(row, col).value
                    if not isinstance(cell, str):
                        string_from_one_row.append(str(cell))
                        break
                    string_from_one_row.append(cell)
                ready_to_use.append(' '.join(string_from_one_row))
            return ready_to_use

    def carry_out_safety_check(self):
        if self.row_start.value() > self.row_end.value():
            self.row_end.setValue(self.row_start.value())
        if self.col_start.value() > self.col_end.value():
            self.col_end.setValue(self.col_start.value())

    def spin_clicked(self):
        if hasattr(self, 'opened_excel'):
            self.carry_out_safety_check()
            self.display_how_many()
            data = self.extract_data_from_pdfs()
            self.data_preview.setText('\n'.join(data))

    def display_how_many(self):
        result = self.row_end.value() - self.row_start.value() + 1
        self.message.setText(f'{str(result)} PDFs will be generated')

    def generate_pdfs(self):
        data = self.extract_data_from_pdfs()
        self.window_signal[list].emit(data)
